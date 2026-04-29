"""
分析问答.xlsx：筛选在8个季度中每季度均有至少1条问答记录的商品。

时间窗口：2022-10-01 ~ 2024-09-30
处理节点：2023-10-01
处理前4个季度：2022-Q4 / 2023-Q1 / 2023-Q2 / 2023-Q3
处理后4个季度：2023-Q4 / 2024-Q1 / 2024-Q2 / 2024-Q3
"""

import openpyxl
from collections import defaultdict
from datetime import datetime


XLSX_PATH = "问答.xlsx"

PRE_QUARTERS = [
    ("pre_Q1", datetime(2022, 10, 1), datetime(2022, 12, 31, 23, 59, 59)),
    ("pre_Q2", datetime(2023,  1, 1), datetime(2023,  3, 31, 23, 59, 59)),
    ("pre_Q3", datetime(2023,  4, 1), datetime(2023,  6, 30, 23, 59, 59)),
    ("pre_Q4", datetime(2023,  7, 1), datetime(2023,  9, 30, 23, 59, 59)),
]
POST_QUARTERS = [
    ("post_Q1", datetime(2023, 10, 1), datetime(2023, 12, 31, 23, 59, 59)),
    ("post_Q2", datetime(2024,  1, 1), datetime(2024,  3, 31, 23, 59, 59)),
    ("post_Q3", datetime(2024,  4, 1), datetime(2024,  6, 30, 23, 59, 59)),
    ("post_Q4", datetime(2024,  7, 1), datetime(2024,  9, 30, 23, 59, 59)),
]
ALL_QUARTERS = PRE_QUARTERS + POST_QUARTERS


def parse_date(value):
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    elif hasattr(value, "year"):
        return value
    return None


def load_quarter_counts(path: str) -> dict[str, dict[str, int]]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    counts: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))

    for row in range(2, ws.max_row + 1):
        auction_id = ws.cell(row=row, column=1).value
        date_val   = ws.cell(row=row, column=2).value
        if not auction_id or not date_val:
            continue
        dt = parse_date(date_val)
        if dt is None:
            continue
        for label, start, end in ALL_QUARTERS:
            if start <= dt <= end:
                counts[auction_id][label] += 1
                break

    return counts


def filter_qualified(counts: dict) -> list[str]:
    return [
        pid for pid, q in counts.items()
        if all(q.get(label, 0) >= 1 for label, _, _ in ALL_QUARTERS)
    ]


def main():
    print(f"读取文件：{XLSX_PATH}")
    counts = load_quarter_counts(XLSX_PATH)
    print(f"数据集商品总数：{len(counts)}")

    qualified = filter_qualified(counts)
    print(f"\n满足条件的商品数：{len(qualified)}")
    print("\n商品ID及各季度记录数：")
    header = "  ".join(label for label, _, _ in ALL_QUARTERS)
    print(f"{'商品ID':<20}  {header}")
    print("-" * 90)
    for pid in sorted(qualified):
        row = "  ".join(
            f"{counts[pid].get(label, 0):>8}"
            for label, _, _ in ALL_QUARTERS
        )
        print(f"{pid:<20}  {row}")


if __name__ == "__main__":
    main()
